from django.shortcuts import render
from . forms import SendMailForm
from django.conf import settings

# Send Email Libraries

import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet

# Create your views here.
def sendmail_view(request):

	form = SendMailForm()
	if request.method == 'POST':
		form = SendMailForm(request.POST, request.FILES)
		if form.is_valid():
			credentials = form.cleaned_data.get('credentials')
			receivers = form.cleaned_data.get('receivers')
			subject = form.cleaned_data.get('subject')
			body = form.cleaned_data.get('body')
			receivers_list = []

			# Credentials File
			cred_workbook = load_workbook(credentials)

			cred_worksheet = cred_workbook['Sheet1']
			row = list(cred_worksheet.rows)[0]

			sender_email = row[0].value
			password = row[1].value

			# Receivers File
			rec_workbook = load_workbook(receivers)

			rec_worksheet = rec_workbook['Sheet1']

			for row in rec_worksheet.rows:
			    if row[0].value != None:
			        receivers_list.append(row[0].value.strip())

			for i in receivers_list:
				print(i)

			for person in receivers_list:

				receiver_email = person
				# Create a multipart message and set headers
				message = MIMEMultipart()
				message["From"] = sender_email
				message["To"] = receiver_email
				message["Subject"] = subject
				message["Bcc"] = receiver_email  # Recommended for mass emails

				# Add body to email
				message.attach(MIMEText(body, "plain"))

				filename = f'{settings.MEDIA_ROOT}/activity.png'

				# Open PDF file in binary mode
				with open(filename, "rb") as attachment:
				    # Add file as application/octet-stream
				    # Email client can usually download this automatically as attachment
				    part = MIMEBase("application", "octet-stream")
				    part.set_payload(attachment.read())

				# Encode file in ASCII characters to send by email    
				encoders.encode_base64(part)

				# Add header as key/value pair to attachment part
				part.add_header(
				    "Content-Disposition",
				    f"attachment; filename= activity.png",
				)

				# Add attachment to message and convert message to string
				message.attach(part)
				text = message.as_string()

				# Log in to server using secure context and send email
				context = ssl.create_default_context()
				with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
				    server.login(sender_email, password)
				    server.sendmail(sender_email, receiver_email, text)

		else:
			print(form.errors)
	context = {}

	return render(request, 'sendmail/sendmail.html', context)
